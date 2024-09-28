import os
from openpyxl import load_workbook
from openpyxl import Workbook

filename='res/local/dblp5-local.xlsx'
if not os.path.exists(filename):
    workbook = Workbook()
    workbook.save(filename)
workbook = load_workbook(filename=filename)
sheet = workbook['Sheet']
mode = "fedavg"
acc = [1,2,3,4]
data = [[mode]] + [[a] for a in acc]
used_columns = sheet.max_column
for row_index, row_data in enumerate(data, start=1):
    for col_index, value in enumerate(row_data, start=used_columns + 1):
        sheet.cell(row=row_index, column=col_index, value=value)
workbook.save(filename=filename)
