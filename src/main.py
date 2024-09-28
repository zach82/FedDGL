# version v2
import os, json
import torch
from fire import Fire
from loguru import logger
from utils import fix_seed
import dataset_loader
from openpyxl import load_workbook, Workbook
import copy

from dataCenter import DataCenter
from Client import Client
import utils as u
import numpy as np
from tqdm.autonotebook import tqdm
from copy import deepcopy
import time
from Server import Server

def main(
    # preprocessing parameters
    time_aggregation=1200000,

    # train / validataion / test ratio
    train_ratio=0.7,
    val_ratio=0.1,

    # augment method
    augment_method='tiara',
    alpha=0.2,
    beta=0.3,
    eps=1e-3,
    K=100,
    symmetric_trick=True,
    dense=False,

    # epochs and early stopping
    epochs=200,
    early_stopping=50,

    # hyperparameters for training
    lr=0.05,
    weight_decay=0.0001,
    lr_decay=0.999,

    # stuffs
    data_dir='data',
    device='cuda:0',
    verbose=False,
    seed=None,

    # model arguments
    model='GCRN',
    input_dim=32,
    output_dim=32,
    decoder_dim=32,

    **kwargs
):
    """
    Main function for experiment

    Parameters
    ----------
    dataset
        dataset name
    time_aggregation
        time step size in secjjonds

    train_ratio
        train split ratio
    val_ratio
        validation split ratio

    augment_method
        dataset augmentation method
    alpha
        teleport probability
    beta
        time treval probability
    eps
        link threshold
    K
        number of diffusion iteration
    symmetric_trick
        method to generate normalized symmetric adjacency matrix
    dense
        whether to use dense adjacency matrix

    epochs
        max epochs
    early_stopping
        early stopping patient
    metric_for_early_stopping
        metric for early stopping

    lr
        learning rate
    weight_decay
        weight decay
    lr_decay
        learning rate decay

    data_dir
        path of the data directory
    device
        device name
    verbose
        verbose mode
    seed
        random seed

    model
        model name
    input_dim
        input feature dimension
    output_dim
        output feature dimension
    decoder_dim
        hidden classifier dimension
    kwargs
        model parameters

    Returns
    -------
    test_metrics
        final test metrics
    history
        training history

    Examples
    --------
    Usage for python code:

    >>> import main as experiment

    Usage for command line:

    $ python src/main.py [--dataset DATASET] [...]
    """
    dataset = kwargs.get('dataset', None)
    split_mode = kwargs.get('split_mode', None)
    mode = kwargs.get('mode', None)
    if dataset in ["DBLP3", "DBLP5"]:
        epochG = 130
    elif dataset in ["Brain", "Reddit"]:
        epochG = 310
    else:
        pass
    if mode in ["fedavgdyn", "fedsagedyn", "fedprotodyn", "mine", "mine_kd", "mine_gp", "mine_pa"]:
        model = "EvolveGCN"
        augment_method = "none"
    elif mode in ["fedavg", "fedsage", "fedproto", "fedego", "fedgcn", "dfedgnn"]:
        model = "GCN"
        augment_method = "none"
    else:
        pass
    if mode in ["mine_kd", "mine_gp", "mine_pa"]:
        base_dir='res/ablation/'
    else:
        base_dir='res/'
    client_num = 10
    model_name = model
    fix_seed(seed)
    torch.cuda.set_device(0)
    dataset_name = dataset
    dataset = dataset_loader.load(
        dataset,
        input_dim,
        train_ratio,
        val_ratio,
        device,
        data_dir,
        seed=seed,
        time_aggregation=time_aggregation
    )
    dataCenter = DataCenter(client_num)
    datas = dataCenter.load_dataSet(dataset_name, dataset, split_mode)
    save_folder_name = kwargs.get('mode', None)
    save_folder_name = f'{save_folder_name}/{dataset_name}/{mode}/{time.time()}/'
    clients = [Client(input_dim, output_dim, device, kwargs, 
                      symmetric_trick, datas[i], alpha, beta, 
                      eps, K, dense, verbose, decoder_dim, epochs, 
                      lr, weight_decay, lr_decay, early_stopping, 
                      augment_method, model, datas[-1], save_folder_name,
                      i) for i in range(client_num)]
    print("clients init finished ", flush=True)
    if mode in ["fedego", "fedproto", "fedprotodyn", "mine", "mine_kd", "mine_gp", "mine_pa"]:
        server = Server(output_dim, device, kwargs, dataset, decoder_dim, 
                 save_folder_name, lr, weight_decay, lr_decay,
                 batch_size=10, margin_threthold=100.0, server_epochs=100)
    train_loss = [0 for i in range(client_num)]
    test_loss = [0 for i in range(client_num)]
    test_metric = [0 for i in range(client_num)]
    testG_loss = [0 for i in range(client_num)]
    testG_metric = [0 for i in range(client_num)]
    test_metric_avgs = []
    testG_metric_avgs = []
    enable_earlystoping = False
    epoch_pbar = tqdm(range(epochG), desc='epoch', position=0)
    best_val_epoch = -1
    best_val_metric = -float('Inf')
    best_model_state = [deepcopy(clients[i].trainer.model.state_dict()) for i in range(client_num)]
    best_decoder_state = [deepcopy(clients[i].trainer.decoder.state_dict()) for i in range(client_num)]
    if mode in ["fedavg", "fedavgdyn", "fedsage", "fedsagedyn"]:
        for epoch in epoch_pbar:
            with torch.no_grad():
                w_avg = u.FedAvg([clients[i].state_dict()
                                for i in range(client_num)])
                for cid in range(client_num):
                    clients[cid].load_state_dict(w_avg)
            for cid in range(client_num):
                train_loss[cid] = clients[cid].supervisedTrain()
            with torch.no_grad():
                for cid in range(client_num):
                    test_loss[cid], test_metric[cid] = clients[cid].test()
                    testG_loss[cid], testG_metric[cid] = clients[cid].testG()
                train_loss_avg = np.mean(train_loss)
                test_metric_avg = np.mean(test_metric)
                testG_metric_avg = np.mean(testG_metric)
            test_metric_avgs.append(test_metric_avg)
            testG_metric_avgs.append(testG_metric_avg)
            epoch_message = 'train loss: {:7.4f}, test metric: {:7.4f}, testG metric: {:7.4f}'.format(
                train_loss_avg, test_metric_avg, testG_metric_avg
            )
            if test_metric_avg > best_val_metric:
                best_val_metric = test_metric_avg
                best_val_epoch = epoch
                best_model_state = [deepcopy(clients[i].trainer.model.state_dict()) for i in range(client_num)]
                best_decoder_state = [deepcopy(clients[i].trainer.decoder.state_dict()) for i in range(client_num)]
            if enable_earlystoping and epoch - best_val_epoch > early_stopping:
                break
            epoch_pbar.write(epoch_message)
        save_acc(base_dir, mode, split_mode, dataset_name, test_metric_avgs, testG_metric_avgs)
        for i in range(client_num):
            clients[i].model.load_state_dict(best_model_state[i])
            clients[i].decoder.load_state_dict(best_decoder_state[i])
            clients[i].trainer.model.eval()
            clients[i].trainer.decoder.eval()
        with torch.no_grad():
            for cid in range(client_num):
                test_loss[cid], test_metric[cid] = clients[cid].test()
                testG_loss[cid], testG_metric[cid] = clients[cid].testG()
        test_metric_avg = np.mean(test_metric)
        testG_metric_avg =np.mean(testG_metric)
        save_lastacc(base_dir, mode, split_mode, dataset_name, test_metric_avg, testG_metric_avg)
        logger.info('dataset: {}'.format(dataset_name))
        logger.info('mode: {}'.format(mode))
        logger.info('seed: {}'.format(seed))
        logger.info('final local metric score is {:7.4f}'.format(test_metric_avg))
        logger.info('final global metric score is {:7.4f}'.format(testG_metric_avg))
        return test_metric_avg, testG_metric_avg
    elif mode in ["fedproto", "fedprotodyn"]:
        _keys = []
        for epoch in epoch_pbar:
            if epoch == 0:
                with torch.no_grad():
                    w_avg = u.FedAvg([clients[i].state_dict()
                                    for i in range(client_num)])
                    for cid in range(client_num):
                        clients[cid].load_state_dict(w_avg, strict=False)
            if epoch == 0:
                for key in w_avg.keys():
                    _keys.append(key)
            for cid in range(client_num):
                train_loss[cid] = clients[cid].supervisedTrain()
            server.receive_protos(clients)
            server.update_Gen()
            classifierWeights = server.calClassifierWeight()
            with torch.no_grad():
                w_avg = u.FedAvg([clients[i].state_dict()
                                for i in range(client_num)])
                for cid in range(client_num):
                    origin_dict = clients[cid].state_dict()
                    deep_dict = {
                        k: (1 - classifierWeights[cid]) * w_avg[k] +
                        classifierWeights[cid] * origin_dict[k]
                        for k in _keys
                    }
                    clients[cid].load_state_dict(deep_dict, strict=False)
            with torch.no_grad():
                for cid in range(client_num):
                    test_loss[cid], test_metric[cid] = clients[cid].test()
                    testG_loss[cid], testG_metric[cid] = clients[cid].testG()
                train_loss_avg = np.mean(train_loss)
                test_metric_avg = np.mean(test_metric)
                testG_metric_avg = np.mean(testG_metric)
            test_metric_avgs.append(test_metric_avg)
            testG_metric_avgs.append(testG_metric_avg)
            epoch_message = 'train loss: {:7.4f}, test metric: {:7.4f}, testG metric: {:7.4f}'.format(
                train_loss_avg, test_metric_avg, testG_metric_avg
            )
            if test_metric_avg > best_val_metric:
                best_val_metric = test_metric_avg
                best_val_epoch = epoch
                best_model_state = [deepcopy(clients[i].trainer.model.state_dict()) for i in range(client_num)]
                best_decoder_state = [deepcopy(clients[i].trainer.decoder.state_dict()) for i in range(client_num)]
            if enable_earlystoping and epoch - best_val_epoch > early_stopping:
                break
            epoch_pbar.write(epoch_message)
        save_acc(base_dir, mode, split_mode, dataset_name, test_metric_avgs, testG_metric_avgs)
        for i in range(client_num):
            clients[i].model.load_state_dict(best_model_state[i])
            clients[i].decoder.load_state_dict(best_decoder_state[i])
            clients[i].trainer.model.eval()
            clients[i].trainer.decoder.eval()
        with torch.no_grad():
            for cid in range(client_num):
                test_loss[cid], test_metric[cid] = clients[cid].test()
                testG_loss[cid], testG_metric[cid] = clients[cid].testG()
        test_metric_avg = np.mean(test_metric)
        testG_metric_avg =np.mean(testG_metric)
        save_lastacc(base_dir, mode, split_mode, dataset_name, test_metric_avg, testG_metric_avg)
        logger.info('dataset: {}'.format(dataset_name))
        logger.info('mode: {}'.format(mode))
        logger.info('seed: {}'.format(seed))
        logger.info('final local metric score is {:7.4f}'.format(test_metric_avg))
        logger.info('final global metric score is {:7.4f}'.format(testG_metric_avg))
        return test_metric_avg, testG_metric_avg
    elif mode in ["fedego"]:
        shallowKeys = []
        deepKeys = []
        for epoch in epoch_pbar:
            with torch.no_grad():
                w_avg = u.FedAvg([clients[i].state_dict()
                                for i in range(client_num)])
                if epoch == 0:
                    for key in w_avg.keys():
                        if (key.startswith("model")):
                            shallowKeys.append(key)
                        else:
                            deepKeys.append(key)
                shallow_dict = {k: w_avg[k] for k in shallowKeys}
                for cid in range(client_num):
                    clients[cid].load_state_dict(shallow_dict, strict=False)
            for cid in range(client_num):
                train_loss[cid] = clients[cid].supervisedTrain()
            server.receive_protos(clients)
            server.update_Gen()
            classifierWeights = server.calClassifierWeight()
            with torch.no_grad():
                server_dict = server.state_dict()
                for cid in range(client_num):
                    origin_dict = clients[cid].state_dict()
                    deep_dict = {
                        k: (1 - classifierWeights[cid]) * server_dict[k] +
                        classifierWeights[cid] * origin_dict[k]
                        for k in deepKeys
                    }
                    clients[cid].load_state_dict(deep_dict, strict=False)
            with torch.no_grad():
                for cid in range(client_num):
                    test_loss[cid], test_metric[cid] = clients[cid].test()
                    testG_loss[cid], testG_metric[cid] = clients[cid].testG()
                train_loss_avg = np.mean(train_loss)
                test_metric_avg = np.mean(test_metric)
                testG_metric_avg = np.mean(testG_metric)
            test_metric_avgs.append(test_metric_avg)
            testG_metric_avgs.append(testG_metric_avg)
            epoch_message = 'train loss: {:7.4f}, test metric: {:7.4f}, testG metric: {:7.4f}'.format(
                train_loss_avg, test_metric_avg, testG_metric_avg
            )
            if test_metric_avg > best_val_metric:
                best_val_metric = test_metric_avg
                best_val_epoch = epoch
                best_model_state = [deepcopy(clients[i].trainer.model.state_dict()) for i in range(client_num)]
                best_decoder_state = [deepcopy(clients[i].trainer.decoder.state_dict()) for i in range(client_num)]
            if enable_earlystoping and epoch - best_val_epoch > early_stopping:
                break
            epoch_pbar.write(epoch_message)
        save_acc(base_dir, mode, split_mode, dataset_name, test_metric_avgs, testG_metric_avgs)
        for i in range(client_num):
            clients[i].model.load_state_dict(best_model_state[i])
            clients[i].decoder.load_state_dict(best_decoder_state[i])
            clients[i].trainer.model.eval()
            clients[i].trainer.decoder.eval()
        with torch.no_grad():
            for cid in range(client_num):
                test_loss[cid], test_metric[cid] = clients[cid].test()
                testG_loss[cid], testG_metric[cid] = clients[cid].testG()
        test_metric_avg = np.mean(test_metric)
        testG_metric_avg =np.mean(testG_metric)
        save_lastacc(base_dir, mode, split_mode, dataset_name, test_metric_avg, testG_metric_avg)
        logger.info('dataset: {}'.format(dataset_name))
        logger.info('mode: {}'.format(mode))
        logger.info('seed: {}'.format(seed))
        logger.info('final local metric score is {:7.4f}'.format(test_metric_avg))
        logger.info('final global metric score is {:7.4f}'.format(testG_metric_avg))
        return test_metric_avg, testG_metric_avg
    elif mode in ["dfedgnn"]:
        for epoch in epoch_pbar:
            for cid in range(client_num):
                train_loss[cid] = clients[cid].supervisedTrain()
            with torch.no_grad():
                w_last = [clients[i].state_dict() for i in range(client_num)]
                w_load = []
                for cid in range(client_num):
                    cur_load = copy.deepcopy(w_last[cid])
                    for k in cur_load.keys():
                        for i in range(-1, 2):
                            if i == 0:
                                continue
                            cur_load[k] += w_last[(cid + i + client_num) %
                                                  client_num][k]
                        cur_load[k] = torch.div(cur_load[k], 3)
                    w_load += [cur_load]
                for cid in range(client_num):
                    clients[cid].load_state_dict(w_load[cid])
                    clients[cid].init_model = w_load[cid]
                for cid in range(client_num):
                    test_loss[cid], test_metric[cid] = clients[cid].test()
                    testG_loss[cid], testG_metric[cid] = clients[cid].testG()
                train_loss_avg = np.mean(train_loss)
                test_metric_avg = np.mean(test_metric)
                testG_metric_avg = np.mean(testG_metric)
            test_metric_avgs.append(test_metric_avg)
            testG_metric_avgs.append(testG_metric_avg)
            epoch_message = 'train loss: {:7.4f}, test metric: {:7.4f}, testG metric: {:7.4f}'.format(
                train_loss_avg, test_metric_avg, testG_metric_avg
            )
            if test_metric_avg > best_val_metric:
                best_val_metric = test_metric_avg
                best_val_epoch = epoch
                best_model_state = [deepcopy(clients[i].trainer.model.state_dict()) for i in range(client_num)]
                best_decoder_state = [deepcopy(clients[i].trainer.decoder.state_dict()) for i in range(client_num)]
            if enable_earlystoping and epoch - best_val_epoch > early_stopping:
                break
            epoch_pbar.write(epoch_message)   
        save_acc(base_dir, mode, split_mode, dataset_name, test_metric_avgs, testG_metric_avgs)
        for i in range(client_num):
            clients[i].model.load_state_dict(best_model_state[i])
            clients[i].decoder.load_state_dict(best_decoder_state[i])
            clients[i].trainer.model.eval()
            clients[i].trainer.decoder.eval()
        with torch.no_grad():
            for cid in range(client_num):
                test_loss[cid], test_metric[cid] = clients[cid].test()
                testG_loss[cid], testG_metric[cid] = clients[cid].testG()
        test_metric_avg = np.mean(test_metric)
        testG_metric_avg =np.mean(testG_metric)
        save_lastacc(base_dir, mode, split_mode, dataset_name, test_metric_avg, testG_metric_avg)
        logger.info('dataset: {}'.format(dataset_name))
        logger.info('mode: {}'.format(mode))
        logger.info('seed: {}'.format(seed))
        logger.info('final local metric score is {:7.4f}'.format(test_metric_avg))
        logger.info('final global metric score is {:7.4f}'.format(testG_metric_avg))
        return test_metric_avg, testG_metric_avg
    elif mode in ["fedgcn"]:
        for epoch in epoch_pbar:
            for cid in range(client_num):
                train_loss[cid] = clients[cid].supervisedTrain()
            with torch.no_grad():
                """load the averaged model"""
                with torch.no_grad():
                    w_avg = u.FedAvg([clients[i].state_dict()
                                      for i in range(client_num)])
                for cid in range(client_num):
                    test_loss[cid], test_metric[cid] = clients[cid].test()
                    testG_loss[cid], testG_metric[cid] = clients[cid].testG()
                train_loss_avg = np.mean(train_loss)
                test_metric_avg = np.mean(test_metric)
                testG_metric_avg = np.mean(testG_metric)
            test_metric_avgs.append(test_metric_avg)
            testG_metric_avgs.append(testG_metric_avg)
            epoch_message = 'train loss: {:7.4f}, test metric: {:7.4f}, testG metric: {:7.4f}'.format(
                train_loss_avg, test_metric_avg, testG_metric_avg
            )
            if test_metric_avg > best_val_metric:
                best_val_metric = test_metric_avg
                best_val_epoch = epoch
                best_model_state = [deepcopy(clients[i].trainer.model.state_dict()) for i in range(client_num)]
                best_decoder_state = [deepcopy(clients[i].trainer.decoder.state_dict()) for i in range(client_num)]
            if enable_earlystoping and epoch - best_val_epoch > early_stopping:
                break
            epoch_pbar.write(epoch_message)
        save_acc(base_dir, mode, split_mode, dataset_name, test_metric_avgs, testG_metric_avgs)
        for i in range(client_num):
            clients[i].model.load_state_dict(best_model_state[i])
            clients[i].decoder.load_state_dict(best_decoder_state[i])
            clients[i].trainer.model.eval()
            clients[i].trainer.decoder.eval()
        with torch.no_grad():
            for cid in range(client_num):
                test_loss[cid], test_metric[cid] = clients[cid].test()
                testG_loss[cid], testG_metric[cid] = clients[cid].testG()
        test_metric_avg = np.mean(test_metric)
        testG_metric_avg =np.mean(testG_metric)
        save_lastacc(base_dir, mode, split_mode, dataset_name, test_metric_avg, testG_metric_avg)
        logger.info('dataset: {}'.format(dataset_name))
        logger.info('mode: {}'.format(mode))
        logger.info('seed: {}'.format(seed))
        logger.info('final local metric score is {:7.4f}'.format(test_metric_avg))
        logger.info('final global metric score is {:7.4f}'.format(testG_metric_avg))
        return test_metric_avg, testG_metric_avg
    elif mode in ["mine"]:
        _keys = []
        for epoch in epoch_pbar:
            if epoch == 0:
                with torch.no_grad():
                    w_avg = u.FedAvg([clients[i].state_dict()
                                    for i in range(client_num)])
                    for cid in range(client_num):
                        clients[cid].load_state_dict(w_avg, strict=False)
            if epoch == 0:
                for key in w_avg.keys():
                    _keys.append(key)
            for cid in range(client_num):
                train_loss[cid] = clients[cid].supervisedTrain()
            server.receive_protos(clients)
            server.update_Gen()
            classifierWeights = server.calClassifierWeight()
            with torch.no_grad():
                w_avg = u.FedAvg([clients[i].state_dict()
                                for i in range(client_num)])
                for cid in range(client_num):
                    origin_dict = clients[cid].state_dict()
                    deep_dict = {
                        k: (1 - classifierWeights[cid]) * w_avg[k] +
                        classifierWeights[cid] * origin_dict[k]
                        for k in _keys
                    }
                    clients[cid].load_state_dict(deep_dict, strict=False)
            with torch.no_grad():
                for cid in range(client_num):
                    test_loss[cid], test_metric[cid] = clients[cid].test()
                    testG_loss[cid], testG_metric[cid] = clients[cid].testG()
                train_loss_avg = np.mean(train_loss)
                test_metric_avg = np.mean(test_metric)
                testG_metric_avg = np.mean(testG_metric)
            test_metric_avgs.append(test_metric_avg)
            testG_metric_avgs.append(testG_metric_avg)
            epoch_message = 'train loss: {:7.4f}, test metric: {:7.4f}, testG metric: {:7.4f}'.format(
                train_loss_avg, test_metric_avg, testG_metric_avg
            )
            if test_metric_avg > best_val_metric:
                best_val_metric = test_metric_avg
                best_val_epoch = epoch
                best_model_state = [deepcopy(clients[i].trainer.model.state_dict()) for i in range(client_num)]
                best_decoder_state = [deepcopy(clients[i].trainer.decoder.state_dict()) for i in range(client_num)]
            if enable_earlystoping and epoch - best_val_epoch > early_stopping:
                break
            epoch_pbar.write(epoch_message)
        save_acc(base_dir, mode, split_mode, dataset_name, test_metric_avgs, testG_metric_avgs)
        for i in range(client_num):
            clients[i].model.load_state_dict(best_model_state[i])
            clients[i].decoder.load_state_dict(best_decoder_state[i])
            clients[i].trainer.model.eval()
            clients[i].trainer.decoder.eval()
        with torch.no_grad():
            for cid in range(client_num):
                test_loss[cid], test_metric[cid] = clients[cid].test()
                testG_loss[cid], testG_metric[cid] = clients[cid].testG()
        test_metric_avg = np.mean(test_metric)
        testG_metric_avg =np.mean(testG_metric)
        save_lastacc(base_dir, mode, split_mode, dataset_name, test_metric_avg, testG_metric_avg)
        logger.info('dataset: {}'.format(dataset_name))
        logger.info('mode: {}'.format(mode))
        logger.info('seed: {}'.format(seed))
        logger.info('final local metric score is {:7.4f}'.format(test_metric_avg))
        logger.info('final global metric score is {:7.4f}'.format(testG_metric_avg))
        return test_metric_avg, testG_metric_avg
    elif mode in ["mine_kd"]:
        _keys = []
        for epoch in epoch_pbar:
            if epoch == 0:
                with torch.no_grad():
                    w_avg = u.FedAvg([clients[i].state_dict()
                                    for i in range(client_num)])
                    for cid in range(client_num):
                        clients[cid].load_state_dict(w_avg, strict=False)
            if epoch == 0:
                for key in w_avg.keys():
                    _keys.append(key)
            for cid in range(client_num):
                train_loss[cid] = clients[cid].supervisedTrain()
            server.receive_protos(clients)
            server.update_Gen()
            classifierWeights = server.calClassifierWeight()
            with torch.no_grad():
                w_avg = u.FedAvg([clients[i].state_dict()
                                for i in range(client_num)])
                for cid in range(client_num):
                    origin_dict = clients[cid].state_dict()
                    deep_dict = {
                        k: (1 - classifierWeights[cid]) * w_avg[k] +
                        classifierWeights[cid] * origin_dict[k]
                        for k in _keys
                    }
                    clients[cid].load_state_dict(deep_dict, strict=False)
            with torch.no_grad():
                for cid in range(client_num):
                    test_loss[cid], test_metric[cid] = clients[cid].test()
                    testG_loss[cid], testG_metric[cid] = clients[cid].testG()
                train_loss_avg = np.mean(train_loss)
                test_metric_avg = np.mean(test_metric)
                testG_metric_avg = np.mean(testG_metric)
            test_metric_avgs.append(test_metric_avg)
            testG_metric_avgs.append(testG_metric_avg)
            epoch_message = 'train loss: {:7.4f}, test metric: {:7.4f}, testG metric: {:7.4f}'.format(
                train_loss_avg, test_metric_avg, testG_metric_avg
            )
            if test_metric_avg > best_val_metric:
                best_val_metric = test_metric_avg
                best_val_epoch = epoch
                best_model_state = [deepcopy(clients[i].trainer.model.state_dict()) for i in range(client_num)]
                best_decoder_state = [deepcopy(clients[i].trainer.decoder.state_dict()) for i in range(client_num)]
            if enable_earlystoping and epoch - best_val_epoch > early_stopping:
                break
            epoch_pbar.write(epoch_message)
        save_acc(base_dir, mode, split_mode, dataset_name, test_metric_avgs, testG_metric_avgs)
        for i in range(client_num):
            clients[i].model.load_state_dict(best_model_state[i])
            clients[i].decoder.load_state_dict(best_decoder_state[i])
            clients[i].trainer.model.eval()
            clients[i].trainer.decoder.eval()
        with torch.no_grad():
            for cid in range(client_num):
                test_loss[cid], test_metric[cid] = clients[cid].test()
                testG_loss[cid], testG_metric[cid] = clients[cid].testG()
        test_metric_avg = np.mean(test_metric)
        testG_metric_avg =np.mean(testG_metric)
        save_lastacc(base_dir, mode, split_mode, dataset_name, test_metric_avg, testG_metric_avg)
        logger.info('dataset: {}'.format(dataset_name))
        logger.info('mode: {}'.format(mode))
        logger.info('seed: {}'.format(seed))
        logger.info('final local metric score is {:7.4f}'.format(test_metric_avg))
        logger.info('final global metric score is {:7.4f}'.format(testG_metric_avg))
        return test_metric_avg, testG_metric_avg
    elif mode in ["mine_gp"]:
        _keys = []
        for epoch in epoch_pbar:
            if epoch == 0:
                with torch.no_grad():
                    w_avg = u.FedAvg([clients[i].state_dict()
                                    for i in range(client_num)])
                    for cid in range(client_num):
                        clients[cid].load_state_dict(w_avg, strict=False)
            if epoch == 0:
                for key in w_avg.keys():
                    _keys.append(key)
            for cid in range(client_num):
                train_loss[cid] = clients[cid].supervisedTrain()
            server.receive_protos(clients)
            server.update_Gen()
            classifierWeights = server.calClassifierWeight()
            with torch.no_grad():
                w_avg = u.FedAvg([clients[i].state_dict()
                                for i in range(client_num)])
                for cid in range(client_num):
                    origin_dict = clients[cid].state_dict()
                    deep_dict = {
                        k: (1 - classifierWeights[cid]) * w_avg[k] +
                        classifierWeights[cid] * origin_dict[k]
                        for k in _keys
                    }
                    clients[cid].load_state_dict(deep_dict, strict=False)
            with torch.no_grad():
                for cid in range(client_num):
                    test_loss[cid], test_metric[cid] = clients[cid].test()
                    testG_loss[cid], testG_metric[cid] = clients[cid].testG()
                train_loss_avg = np.mean(train_loss)
                test_metric_avg = np.mean(test_metric)
                testG_metric_avg = np.mean(testG_metric)
            test_metric_avgs.append(test_metric_avg)
            testG_metric_avgs.append(testG_metric_avg)
            epoch_message = 'train loss: {:7.4f}, test metric: {:7.4f}, testG metric: {:7.4f}'.format(
                train_loss_avg, test_metric_avg, testG_metric_avg
            )
            if test_metric_avg > best_val_metric:
                best_val_metric = test_metric_avg
                best_val_epoch = epoch
                best_model_state = [deepcopy(clients[i].trainer.model.state_dict()) for i in range(client_num)]
                best_decoder_state = [deepcopy(clients[i].trainer.decoder.state_dict()) for i in range(client_num)]
            if enable_earlystoping and epoch - best_val_epoch > early_stopping:
                break
            epoch_pbar.write(epoch_message)
        save_acc(base_dir, mode, split_mode, dataset_name, test_metric_avgs, testG_metric_avgs)
        for i in range(client_num):
            clients[i].model.load_state_dict(best_model_state[i])
            clients[i].decoder.load_state_dict(best_decoder_state[i])
            clients[i].trainer.model.eval()
            clients[i].trainer.decoder.eval()
        with torch.no_grad():
            for cid in range(client_num):
                test_loss[cid], test_metric[cid] = clients[cid].test()
                testG_loss[cid], testG_metric[cid] = clients[cid].testG()
        test_metric_avg = np.mean(test_metric)
        testG_metric_avg =np.mean(testG_metric)
        save_lastacc(base_dir, mode, split_mode, dataset_name, test_metric_avg, testG_metric_avg)
        logger.info('dataset: {}'.format(dataset_name))
        logger.info('mode: {}'.format(mode))
        logger.info('seed: {}'.format(seed))
        logger.info('final local metric score is {:7.4f}'.format(test_metric_avg))
        logger.info('final global metric score is {:7.4f}'.format(testG_metric_avg))
        return test_metric_avg, testG_metric_avg
    elif mode in ["mine_pa"]:
        _keys = []
        for epoch in epoch_pbar:
            if epoch == 0:
                with torch.no_grad():
                    w_avg = u.FedAvg([clients[i].state_dict()
                                    for i in range(client_num)])
                    for cid in range(client_num):
                        clients[cid].load_state_dict(w_avg, strict=False)
            if epoch == 0:
                for key in w_avg.keys():
                    _keys.append(key)
            for cid in range(client_num):
                train_loss[cid] = clients[cid].supervisedTrain()
            server.receive_protos(clients)
            server.update_Gen()
            classifierWeights = server.calClassifierWeight()
            with torch.no_grad():
                w_avg = u.FedAvg([clients[i].state_dict()
                                for i in range(client_num)])
                for cid in range(client_num):
                    origin_dict = clients[cid].state_dict()
                    deep_dict = {
                        k: (1 - classifierWeights[cid]) * w_avg[k] +
                        classifierWeights[cid] * origin_dict[k]
                        for k in _keys
                    }
                    clients[cid].load_state_dict(deep_dict, strict=False)
            with torch.no_grad():
                for cid in range(client_num):
                    test_loss[cid], test_metric[cid] = clients[cid].test()
                    testG_loss[cid], testG_metric[cid] = clients[cid].testG()
                train_loss_avg = np.mean(train_loss)
                test_metric_avg = np.mean(test_metric)
                testG_metric_avg = np.mean(testG_metric)
            test_metric_avgs.append(test_metric_avg)
            testG_metric_avgs.append(testG_metric_avg)
            epoch_message = 'train loss: {:7.4f}, test metric: {:7.4f}, testG metric: {:7.4f}'.format(
                train_loss_avg, test_metric_avg, testG_metric_avg
            )
            if test_metric_avg > best_val_metric:
                best_val_metric = test_metric_avg
                best_val_epoch = epoch
                best_model_state = [deepcopy(clients[i].trainer.model.state_dict()) for i in range(client_num)]
                best_decoder_state = [deepcopy(clients[i].trainer.decoder.state_dict()) for i in range(client_num)]
            if enable_earlystoping and epoch - best_val_epoch > early_stopping:
                break
            epoch_pbar.write(epoch_message)
        save_acc(base_dir, mode, split_mode, dataset_name, test_metric_avgs, testG_metric_avgs)
        for i in range(client_num):
            clients[i].model.load_state_dict(best_model_state[i])
            clients[i].decoder.load_state_dict(best_decoder_state[i])
            clients[i].trainer.model.eval()
            clients[i].trainer.decoder.eval()
        with torch.no_grad():
            for cid in range(client_num):
                test_loss[cid], test_metric[cid] = clients[cid].test()
                testG_loss[cid], testG_metric[cid] = clients[cid].testG()
        test_metric_avg = np.mean(test_metric)
        testG_metric_avg =np.mean(testG_metric)
        save_lastacc(base_dir, mode, split_mode, dataset_name, test_metric_avg, testG_metric_avg)
        logger.info('dataset: {}'.format(dataset_name))
        logger.info('mode: {}'.format(mode))
        logger.info('seed: {}'.format(seed))
        logger.info('final local metric score is {:7.4f}'.format(test_metric_avg))
        logger.info('final global metric score is {:7.4f}'.format(testG_metric_avg))
        return test_metric_avg, testG_metric_avg
    else:
        pass
def print_acc(test_metric_avgs):
    for test_metric_avg in test_metric_avgs:
        print(test_metric_avg)
def save_acc(base_dir, mode, split_mode, dataset, test_metric_avgs, testG_metric_avgs):
    filename='{}{}-local/{}-local.xlsx'.format(base_dir, split_mode, dataset)
    if not os.path.exists(filename):
        workbook = Workbook()
        workbook.save(filename)
    workbook = load_workbook(filename=filename)
    sheet = workbook['Sheet']
    data = [[mode]] + [[a] for a in test_metric_avgs]
    used_columns = sheet.max_column
    for row_index, row_data in enumerate(data, start=1):
        for col_index, value in enumerate(row_data, start=used_columns + 1):
            sheet.cell(row=row_index, column=col_index, value=value)
    workbook.save(filename=filename)
    filename='{}{}-global/{}-global.xlsx'.format(base_dir, split_mode, dataset)
    if not os.path.exists(filename):
        workbook = Workbook()
        workbook.save(filename)
    workbook = load_workbook(filename=filename)
    sheet = workbook['Sheet']
    data = [[mode]] + [[a] for a in testG_metric_avgs]
    used_columns = sheet.max_column
    for row_index, row_data in enumerate(data, start=1):
        for col_index, value in enumerate(row_data, start=used_columns + 1):
            sheet.cell(row=row_index, column=col_index, value=value)
    workbook.save(filename=filename)
def save_lastacc(base_dir, mode, split_mode, dataset, test_metric_avg, testG_metric_avg):
    filename='{}{}-locallast/{}-local.xlsx'.format(base_dir, split_mode, dataset)
    if not os.path.exists(filename):
        workbook = Workbook()
        workbook.save(filename)
    workbook = load_workbook(filename=filename)
    sheet = workbook['Sheet']
    data = [[mode]] + [[test_metric_avg]]
    used_columns = sheet.max_column
    for row_index, row_data in enumerate(data, start=1):
        for col_index, value in enumerate(row_data, start=used_columns + 1):
            sheet.cell(row=row_index, column=col_index, value=value)
    workbook.save(filename=filename)
    filename='{}{}-globallast/{}-global.xlsx'.format(base_dir, split_mode, dataset)
    if not os.path.exists(filename):
        workbook = Workbook()
        workbook.save(filename)
    workbook = load_workbook(filename=filename)
    sheet = workbook['Sheet']
    data = [[mode]] + [[testG_metric_avg]]
    used_columns = sheet.max_column
    for row_index, row_data in enumerate(data, start=1):
        for col_index, value in enumerate(row_data, start=used_columns + 1):
            sheet.cell(row=row_index, column=col_index, value=value)
    workbook.save(filename=filename)
def main_wraper(**kwargs):
    conf_file = kwargs.get('conf_file', None)
    if conf_file is not None:
        config = json.load(open(conf_file, 'r'))
        for k in kwargs:
            if k in config:
                logger.warning('{} will be overwritten!'.format(k))
        kwargs = {**kwargs, **{ "test":"local" }}
        kwargs = {**kwargs, **{ "save_folder_name":"temp" }}
        kwargs = {**kwargs, **{ "dataset":config['dataset'] }}
        setting = {**config, **kwargs}
        report_setting(setting)
        test_metric = main(**setting)
    else:
        report_setting(kwargs)
        main(**kwargs)
def report_setting(setting):
    for k, v in setting.items():
        logger.info('{}: {}'.format(k, v))
if __name__ == "__main__":
    Fire(main_wraper)
